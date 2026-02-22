"""
Router: reports.py
- GET /api/reports/serials/export        → Excel export of all serials
- GET /api/reports/serials/export-csv   → CSV export
- GET /api/reports/warranty/{serial}    → PDF warranty certificate (download)
- POST /api/reports/serials/import      → Bulk serial import via uploaded CSV
"""

import csv
import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import get_db
from models import Serial, Product, User
from auth import get_current_admin

router = APIRouter(prefix="/api/reports", tags=["reports"])


# ─── Excel Export ─────────────────────────────────────────────────────────────

@router.get("/serials/export")
def export_serials_excel(db: Session = Depends(get_db), _=Depends(get_current_admin)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    serials = db.query(Serial).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Serials"

    headers = ["الرقم التسلسلي", "المنتج", "الحالة", "العميل", "الهاتف", "تاريخ الشراء", "تاريخ الإنشاء"]
    # Style header row
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    status_map = {"active": "نشط", "expired": "منتهي", "inactive": "غير مفعل", "void": "ملغي"}

    for row_i, s in enumerate(serials, 2):
        product_name = s.product.name if s.product else ""
        customer_name = s.user.name if s.user else ""
        customer_phone = s.user.phone if s.user else ""
        ws.append([
            s.serial_number,
            product_name,
            status_map.get(s.warranty_status, s.warranty_status),
            customer_name,
            customer_phone,
            str(s.purchase_date) if s.purchase_date else "",
            str(s.created_at)[:10] if s.created_at else "",
        ])
        if row_i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_i, column=col).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=serials_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


# ─── PDF Warranty Certificate ─────────────────────────────────────────────────

@router.get("/warranty/{serial_number}/pdf")
def download_warranty_pdf(serial_number: str, db: Session = Depends(get_db)):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        raise HTTPException(500, "reportlab not installed")

    s = db.query(Serial).filter(Serial.serial_number == serial_number).first()
    if not s:
        raise HTTPException(404, "Serial not found")

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    # Background
    c.setFillColor(colors.HexColor("#6366F1"))
    c.rect(0, h - 120, w, 120, fill=1, stroke=0)

    # Title
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 24)
    c.drawCentredString(w / 2, h - 55, "Smart Warranty Certificate")
    c.setFont("Helvetica", 13)
    c.drawCentredString(w / 2, h - 80, "شهادة الضمان الرسمية")

    # Content box
    c.setFillColor(colors.HexColor("#F8FAFC"))
    c.roundRect(40, h - 480, w - 80, 340, 12, fill=1, stroke=0)

    # Fields
    def field(label_en: str, label_ar: str, value: str, y: int):
        c.setFillColor(colors.HexColor("#64748B"))
        c.setFont("Helvetica", 10)
        c.drawString(60, y + 16, label_en)
        c.setFillColor(colors.HexColor("#1E293B"))
        c.setFont("Helvetica-Bold", 13)
        c.drawString(60, y, value or "N/A")

    product_name = s.product.name if s.product else "N/A"
    customer_name = s.user.name if s.user else "N/A"
    purchase_date = str(s.purchase_date) if s.purchase_date else "N/A"
    warranty_months = str(s.product.warranty_months) + " months" if s.product else "N/A"
    status_map = {"active": "Active ✓", "expired": "Expired ✗", "inactive": "Not Activated", "void": "Void"}
    status = status_map.get(s.warranty_status, s.warranty_status)

    field("Product", "المنتج", product_name, h - 190)
    field("Serial Number", "الرقم التسلسلي", serial_number, h - 240)
    field("Customer", "العميل", customer_name, h - 290)
    field("Purchase Date", "تاريخ الشراء", purchase_date, h - 340)
    field("Warranty Period", "مدة الضمان", warranty_months, h - 390)

    # Status badge
    status_color = colors.HexColor("#10B981") if s.warranty_status == "active" else colors.HexColor("#EF4444")
    c.setFillColor(status_color)
    c.roundRect(60, h - 455, 160, 30, 8, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(140, h - 436, status)

    # Footer
    c.setFillColor(colors.HexColor("#94A3B8"))
    c.setFont("Helvetica", 9)
    c.drawCentredString(w / 2, 40, f"Generated: {datetime.now().strftime('%Y-%m-%d')}  |  Smart Warranty Tracker")

    c.save()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=warranty_{serial_number}.pdf"}
    )


# ─── Bulk CSV Import ───────────────────────────────────────────────────────────

@router.post("/serials/import")
async def bulk_import_serials(
    file: UploadFile = File(...),
    product_id: int = 0,
    db: Session = Depends(get_db),
    _=Depends(get_current_admin)
):
    """
    Upload a CSV with columns: serial_number (or auto-generated if empty), product_id (optional)
    Returns: count of imported serials and list of duplicates skipped.
    """
    if not file.filename.endswith(".csv"):
        raise HTTPException(400, "Only CSV files are accepted")

    content = await file.read()
    text = content.decode("utf-8-sig")  # Handle BOM
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    skipped = []

    for row in reader:
        serial_num = row.get("serial_number", "").strip().upper()
        pid = int(row.get("product_id", product_id) or product_id)

        if not serial_num:
            continue
        if not pid:
            continue

        product = db.query(Product).filter(Product.id == pid).first()
        if not product:
            skipped.append(f"{serial_num} (product not found)")
            continue

        exists = db.query(Serial).filter(Serial.serial_number == serial_num).first()
        if exists:
            skipped.append(f"{serial_num} (duplicate)")
            continue

        db.add(Serial(serial_number=serial_num, product_id=pid))
        imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped, "message": f"تم استيراد {imported} سيريال بنجاح"}
